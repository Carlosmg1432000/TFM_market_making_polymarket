import json
import sqlite3
from pathlib import Path

import numpy as np
import pandas as pd
import requests


# ============================================================
# CONFIGURACIÓN
# ============================================================

DB_PATH = Path("data") / "market_sim.db"

TABLE_NAME = "orderbook_snapshots"

OUTPUT_DIR = Path("outputs")

CSV_OUTPUT = OUTPUT_DIR / "sample_summary.csv"

XLSX_OUTPUT = OUTPUT_DIR / "sample_summary.xlsx"

GAMMA_MARKETS_URL = "https://gamma-api.polymarket.com/markets"

REQUEST_TIMEOUT = 15


# ============================================================
# FUNCIONES AUXILIARES SQLITE
# ============================================================

def table_exists(cursor, table_name):
    """
    Comprueba si existe una tabla en SQLite.
    """
    cursor.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type='table' AND name=?;
        """,
        (table_name,)
    )
    return cursor.fetchone() is not None


def load_snapshots():
    """
    Carga todos los snapshots guardados en SQLite.
    """
    if not DB_PATH.exists():
        print(f"No existe la base de datos: {DB_PATH.resolve()}")
        return pd.DataFrame()

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    if not table_exists(cursor, TABLE_NAME):
        print(f"No existe la tabla {TABLE_NAME} en la base de datos.")
        conn.close()
        return pd.DataFrame()

    query = f"""
        SELECT
            id,
            timestamp,
            token_id,
            best_bid,
            best_ask,
            midprice,
            spread,
            bid_depth,
            ask_depth,
            total_depth,
            imbalance,
            top_bid_size,
            top_ask_size,
            book_volume_proxy
        FROM {TABLE_NAME}
        ORDER BY timestamp ASC;
    """

    df = pd.read_sql_query(query, conn)
    conn.close()

    if df.empty:
        print("La tabla existe, pero no contiene snapshots.")
        return df

    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df["token_id"] = df["token_id"].astype(str)

    numeric_cols = [
        "id",
        "best_bid",
        "best_ask",
        "midprice",
        "spread",
        "bid_depth",
        "ask_depth",
        "total_depth",
        "imbalance",
        "top_bid_size",
        "top_ask_size",
        "book_volume_proxy",
    ]

    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.dropna(subset=["timestamp", "token_id", "midprice", "spread"])

    return df


# ============================================================
# FUNCIONES GAMMA API / METADATOS DEL MERCADO
# ============================================================

def safe_json_loads(value):
    """
    Convierte campos que a veces vienen como string JSON.
    Si ya vienen como lista/dict, los devuelve directamente.
    """
    if value is None:
        return None

    if isinstance(value, (list, dict)):
        return value

    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return value

    return value


def normalize_list_field(value):
    """
    Normaliza un campo que puede venir como lista o como string JSON.
    """
    parsed = safe_json_loads(value)

    if parsed is None:
        return []

    if isinstance(parsed, list):
        return [str(x) for x in parsed]

    if isinstance(parsed, str):
        return [parsed]

    return []


def get_market_metadata_from_gamma(token_id):
    """
    Busca en Gamma API el mercado asociado a un token_id de CLOB.

    Devuelve un diccionario con:
    - market_title
    - event_title
    - question
    - outcome
    - market_slug
    - event_slug
    - volume
    - volume_24hr
    - liquidity
    - end_date
    """
    empty_result = {
        "market_title": "No localizado",
        "event_title": "No localizado",
        "question": "No localizado",
        "outcome": "No localizado",
        "market_slug": "No localizado",
        "event_slug": "No localizado",
        "volume": np.nan,
        "volume_24hr": np.nan,
        "liquidity": np.nan,
        "end_date": "No localizado",
    }

    token_id = str(token_id)

    try:
        response = requests.get(
            GAMMA_MARKETS_URL,
            params={
                "clob_token_ids": token_id,
                "limit": 20,
            },
            timeout=REQUEST_TIMEOUT,
            headers={
                "User-Agent": "Mozilla/5.0",
                "Accept": "application/json",
            }
        )

        response.raise_for_status()
        data = response.json()

        if isinstance(data, dict):
            markets = data.get("markets") or data.get("data") or []
        elif isinstance(data, list):
            markets = data
        else:
            markets = []

        if not markets:
            return empty_result

        for market in markets:
            clob_token_ids = normalize_list_field(
                market.get("clobTokenIds") or market.get("clob_token_ids")
            )

            if token_id not in clob_token_ids:
                continue

            outcomes = normalize_list_field(market.get("outcomes"))

            outcome = "No localizado"

            try:
                idx = clob_token_ids.index(token_id)
                if idx < len(outcomes):
                    outcome = outcomes[idx]
            except Exception:
                outcome = "No localizado"

            event_title = "No localizado"
            event_slug = "No localizado"

            events = market.get("events")

            if isinstance(events, list) and len(events) > 0:
                first_event = events[0]
                if isinstance(first_event, dict):
                    event_title = (
                        first_event.get("title")
                        or first_event.get("name")
                        or first_event.get("ticker")
                        or "No localizado"
                    )
                    event_slug = first_event.get("slug") or "No localizado"

            if event_title == "No localizado":
                event_title = (
                    market.get("eventTitle")
                    or market.get("event_title")
                    or market.get("groupItemTitle")
                    or market.get("title")
                    or market.get("question")
                    or "No localizado"
                )

            result = {
                "market_title": (
                    market.get("title")
                    or market.get("question")
                    or market.get("description")
                    or "No localizado"
                ),
                "event_title": event_title,
                "question": market.get("question") or market.get("title") or "No localizado",
                "outcome": outcome,
                "market_slug": market.get("slug") or "No localizado",
                "event_slug": event_slug,
                "volume": to_float_or_nan(
                    market.get("volume")
                    or market.get("volumeNum")
                    or market.get("volume_num")
                ),
                "volume_24hr": to_float_or_nan(
                    market.get("volume24hr")
                    or market.get("volume_24hr")
                    or market.get("volume24Hr")
                    or market.get("volume24hrNum")
                ),
                "liquidity": to_float_or_nan(
                    market.get("liquidity")
                    or market.get("liquidityNum")
                    or market.get("liquidity_num")
                ),
                "end_date": (
                    market.get("endDate")
                    or market.get("end_date")
                    or market.get("endDateIso")
                    or "No localizado"
                ),
            }

            return result

        return empty_result

    except Exception as e:
        print(f"No se pudo recuperar metadata para token_id {token_id}. Error: {e}")
        return empty_result


def to_float_or_nan(value):
    """
    Convierte un valor a float si es posible.
    """
    try:
        if value is None:
            return np.nan
        return float(value)
    except Exception:
        return np.nan


def build_metadata_table(token_ids):
    """
    Crea una tabla con los metadatos de mercado para cada token_id.
    """
    rows = []

    print("\nBuscando nombres de mercado en Gamma API...")

    for i, token_id in enumerate(token_ids, start=1):
        print(f"{i}/{len(token_ids)} buscando token_id: {token_id}")

        metadata = get_market_metadata_from_gamma(token_id)
        metadata["token_id"] = str(token_id)

        rows.append(metadata)

    metadata_df = pd.DataFrame(rows)

    ordered_cols = [
        "token_id",
        "market_title",
        "event_title",
        "question",
        "outcome",
        "market_slug",
        "event_slug",
        "volume",
        "volume_24hr",
        "liquidity",
        "end_date",
    ]

    for col in ordered_cols:
        if col not in metadata_df.columns:
            metadata_df[col] = np.nan

    metadata_df = metadata_df[ordered_cols]

    return metadata_df


# ============================================================
# RESUMEN DE MUESTRAS
# ============================================================

def summarize_one_sample(group):
    """
    Calcula las métricas principales de una muestra/token_id.
    """
    group = group.sort_values("timestamp").copy()

    mid_diff = group["midprice"].diff()
    spread_diff = group["spread"].diff()
    imbalance_diff = group["imbalance"].diff()

    n_snapshots = len(group)

    start_time = group["timestamp"].min()
    end_time = group["timestamp"].max()

    duration_minutes = (end_time - start_time).total_seconds() / 60

    mid_unique = group["midprice"].nunique()
    spread_unique = group["spread"].nunique()

    mid_changes = int((mid_diff.fillna(0) != 0).sum())
    spread_changes = int((spread_diff.fillna(0) != 0).sum())
    imbalance_changes = int((imbalance_diff.fillna(0) != 0).sum())

    if n_snapshots > 1:
        mid_change_rate = mid_changes / (n_snapshots - 1)
        spread_change_rate = spread_changes / (n_snapshots - 1)
        imbalance_change_rate = imbalance_changes / (n_snapshots - 1)
    else:
        mid_change_rate = 0
        spread_change_rate = 0
        imbalance_change_rate = 0

    summary = {
        "token_id": group["token_id"].iloc[0],
        "n_snapshots": n_snapshots,
        "start_time": start_time,
        "end_time": end_time,
        "duration_minutes": duration_minutes,

        "mid_min": group["midprice"].min(),
        "mid_max": group["midprice"].max(),
        "mid_mean": group["midprice"].mean(),
        "mid_std": group["midprice"].std(),
        "mid_unique": mid_unique,
        "mid_changes": mid_changes,
        "mid_change_rate": mid_change_rate,

        "spread_min": group["spread"].min(),
        "spread_max": group["spread"].max(),
        "spread_mean": group["spread"].mean(),
        "spread_median": group["spread"].median(),
        "spread_std": group["spread"].std(),
        "spread_unique": spread_unique,
        "spread_changes": spread_changes,
        "spread_change_rate": spread_change_rate,

        "total_depth_mean": group["total_depth"].mean(),
        "total_depth_median": group["total_depth"].median(),
        "total_depth_min": group["total_depth"].min(),
        "total_depth_max": group["total_depth"].max(),
        "total_depth_std": group["total_depth"].std(),

        "bid_depth_mean": group["bid_depth"].mean(),
        "ask_depth_mean": group["ask_depth"].mean(),

        "imbalance_mean": group["imbalance"].mean(),
        "imbalance_std": group["imbalance"].std(),
        "imbalance_min": group["imbalance"].min(),
        "imbalance_max": group["imbalance"].max(),
        "imbalance_abs_mean": group["imbalance"].abs().mean(),
        "imbalance_changes": imbalance_changes,
        "imbalance_change_rate": imbalance_change_rate,

        "book_volume_proxy_mean": group["book_volume_proxy"].mean(),
        "book_volume_proxy_median": group["book_volume_proxy"].median(),
        "book_volume_proxy_min": group["book_volume_proxy"].min(),
        "book_volume_proxy_max": group["book_volume_proxy"].max(),
    }

    return pd.Series(summary)


def add_selection_score(summary_df):
    """
    Añade una puntuación orientativa para escoger la muestra más útil.

    Criterios:
    - número de snapshots,
    - cambios del midprice,
    - variabilidad del midprice,
    - variabilidad del imbalance,
    - profundidad visible media.
    """
    df = summary_df.copy()

    score_cols = {
        "n_snapshots": 0.20,
        "mid_changes": 0.30,
        "mid_std": 0.20,
        "imbalance_std": 0.15,
        "total_depth_mean": 0.15,
    }

    df["selection_score"] = 0.0

    for col, weight in score_cols.items():
        if col in df.columns:
            values = df[col].replace([np.inf, -np.inf], np.nan).fillna(0)

            if values.nunique() <= 1:
                normalized = pd.Series(0.0, index=df.index)
            else:
                normalized = (values - values.min()) / (values.max() - values.min())

            df["selection_score"] += weight * normalized

    df = df.sort_values("selection_score", ascending=False).reset_index(drop=True)

    return df


def build_sample_summary():
    """
    Genera la tabla resumen de todas las muestras disponibles.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    df = load_snapshots()

    if df.empty:
        print("No hay datos suficientes para generar la tabla de muestras.")
        return pd.DataFrame()

    print(f"Snapshots cargados desde SQLite: {len(df)}")
    print(f"Número de token_id distintos: {df['token_id'].nunique()}")

    summary_rows = []

    for token_id, group in df.groupby("token_id"):
        summary = summarize_one_sample(group)
        summary_rows.append(summary)

    summary_df = pd.DataFrame(summary_rows)

    unique_token_ids = summary_df["token_id"].astype(str).unique().tolist()

    metadata_df = build_metadata_table(unique_token_ids)

    summary_df["token_id"] = summary_df["token_id"].astype(str)
    metadata_df["token_id"] = metadata_df["token_id"].astype(str)

    summary_df = summary_df.merge(
        metadata_df,
        on="token_id",
        how="left"
    )

    summary_df = add_selection_score(summary_df)

    first_cols = [
        "market_title",
        "event_title",
        "question",
        "outcome",
        "token_id",
        "n_snapshots",
        "duration_minutes",
        "mid_changes",
        "mid_unique",
        "mid_std",
        "spread_mean",
        "total_depth_mean",
        "imbalance_std",
        "selection_score",
        "volume",
        "volume_24hr",
        "liquidity",
        "end_date",
    ]

    existing_first_cols = [col for col in first_cols if col in summary_df.columns]
    remaining_cols = [col for col in summary_df.columns if col not in existing_first_cols]

    summary_df = summary_df[existing_first_cols + remaining_cols]

    summary_df.to_csv(CSV_OUTPUT, index=False)

    print(f"Archivo CSV generado: {CSV_OUTPUT.resolve()}")

    save_excel(summary_df, df, metadata_df)

    print("\nResumen de muestras:")
    print(summary_df)

    if not summary_df.empty:
        print("\nMejor muestra orientativa según selection_score:")
        best = summary_df.iloc[0]
        print(f"market_title: {best.get('market_title')}")
        print(f"event_title: {best.get('event_title')}")
        print(f"question: {best.get('question')}")
        print(f"outcome: {best.get('outcome')}")
        print(f"token_id: {best.get('token_id')}")
        print(f"selection_score: {best.get('selection_score'):.4f}")
        print(f"n_snapshots: {best.get('n_snapshots')}")
        print(f"mid_changes: {best.get('mid_changes')}")
        print(f"mid_std: {best.get('mid_std')}")
        print(f"spread_mean: {best.get('spread_mean')}")
        print(f"total_depth_mean: {best.get('total_depth_mean')}")
        print(f"imbalance_std: {best.get('imbalance_std')}")

    return summary_df


# ============================================================
# GUARDADO EN EXCEL
# ============================================================

def save_excel(summary_df, raw_df, metadata_df):
    """
    Guarda el resumen en Excel.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("No se pudo generar Excel. Falta instalar openpyxl.")
        print("Ejecuta: pip install openpyxl")
        return

    wb = Workbook()

    # --------------------------------------------------------
    # Hoja 1: resumen de muestras
    # --------------------------------------------------------
    ws = wb.active
    ws.title = "Resumen muestras"

    write_dataframe_to_sheet(ws, summary_df)

    # --------------------------------------------------------
    # Hoja 2: mejor muestra
    # --------------------------------------------------------
    ws_best = wb.create_sheet("Mejor muestra")

    if not summary_df.empty:
        best = summary_df.iloc[0]

        best_rows = [
            ("market_title", best.get("market_title")),
            ("event_title", best.get("event_title")),
            ("question", best.get("question")),
            ("outcome", best.get("outcome")),
            ("token_id", best.get("token_id")),
            ("selection_score", best.get("selection_score")),
            ("n_snapshots", best.get("n_snapshots")),
            ("duration_minutes", best.get("duration_minutes")),
            ("mid_changes", best.get("mid_changes")),
            ("mid_unique", best.get("mid_unique")),
            ("mid_std", best.get("mid_std")),
            ("spread_mean", best.get("spread_mean")),
            ("total_depth_mean", best.get("total_depth_mean")),
            ("imbalance_std", best.get("imbalance_std")),
            ("volume", best.get("volume")),
            ("volume_24hr", best.get("volume_24hr")),
            ("liquidity", best.get("liquidity")),
            ("end_date", best.get("end_date")),
        ]

        best_df = pd.DataFrame(best_rows, columns=["Métrica", "Valor"])
        write_dataframe_to_sheet(ws_best, best_df)

    # --------------------------------------------------------
    # Hoja 3: metadata mercados
    # --------------------------------------------------------
    ws_metadata = wb.create_sheet("Metadata mercados")
    write_dataframe_to_sheet(ws_metadata, metadata_df)

    # --------------------------------------------------------
    # Hoja 4: snapshots crudos
    # --------------------------------------------------------
    ws_raw = wb.create_sheet("Snapshots crudos")
    write_dataframe_to_sheet(ws_raw, raw_df)

    wb.save(XLSX_OUTPUT)

    print(f"Archivo Excel generado correctamente: {XLSX_OUTPUT.resolve()}")


def write_dataframe_to_sheet(ws, df):
    """
    Escribe un DataFrame en una hoja de Excel usando openpyxl.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment

    if df is None or df.empty:
        ws.cell(row=1, column=1, value="Sin datos")
        return

    # Cabeceras
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(df.columns, start=1):
            value = row[col_name]

            if pd.isna(value):
                value = None

            if isinstance(value, pd.Timestamp):
                value = value.strftime("%Y-%m-%d %H:%M:%S")

            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))

        for row_idx in range(2, len(df) + 2):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        adjusted_width = min(max_len + 2, 70)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width


if __name__ == "__main__":
    build_sample_summary()